import util
import requests
import re
import openpyxl
import cardsdb
from openpyxl.utils.cell import get_column_letter
from langconv import Converter

def translate_online(name):
    if name == 'Card name':
        return '卡牌名称'
    content = requests.get(util.link_format(name=name)).text
    pattern = re.compile('<a href="/card/(.*?)" class="el-card-visual-content">', re.S)
    results = re.findall(pattern, content)
    if len(results) == 0:
        return name
    else:
        return util.get_card_data(int(results[0]))['title']
    
def translate_local(cursor, name):
    if name == 'Card name':
        return '卡牌名称'
    elif name == 'Ratio':
        return '比率'
    elif name == 'W.Av.':
        return '胜率加权平均采用数'
    elif name == 'S.Av.':
        return '平均采用数'
    elif name == None:
        return None
    else:
        exact = 0
        if len(str(name)) <= 5:
            exact = 1
        return cardsdb.translate(cursor, str(name), lang1=1, lang2=3, exact=exact)[0]

def find_base_cell(sheet):
    for i in range(1,6):
        for j in range(1,6):
            if sheet.cell(i, j).value == 'Card name':
                return (i, j)
    return (1,1)

if __name__ == '__main__':
    (conn, cursor) = cardsdb.init()
    filename = '大表'
    wb = openpyxl.load_workbook(filename + '.xlsx')
    sheet = wb[wb.sheetnames[0]]
    (base_row, base_col) = find_base_cell(sheet)
    for r in range(base_row, sheet.max_row + 1):
        cell = sheet.cell(r, base_col).value
        if cell == 'Card name':
            for i in range(4):
                sheet.cell(r, base_col + i).value = translate_local(cursor, sheet.cell(r, base_col + i).value)
        else:
            sheet.cell(r, base_col).value = translate_local(cursor, cell)
            for i in range(1, 4):
                sheet.cell(r, base_col + i).number_format = '0.0000'
    sheet.column_dimensions[get_column_letter(base_col)].width = 13
    sheet.column_dimensions[get_column_letter(base_col+1)].width = 8
    sheet.column_dimensions[get_column_letter(base_col+2)].width = 11
    sheet.column_dimensions[get_column_letter(base_col+3)].width = 11
    wb.save(filename + '翻译版.xlsx')
    
    for r in range(base_row, sheet.max_row + 1):
        cell = sheet.cell(r, base_col).value
        if cell != None:
            sheet.cell(r, base_col).value = Converter('zh-hans').convert(cell)
    wb.save(filename + '简体版.xlsx')
    conn.close()