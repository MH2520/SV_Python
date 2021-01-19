import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.cell import get_column_letter
import numpy as np
from deckname_dict import deckname_dict
from deckmerge_dict import deckmerge_dict
rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFF8696B',
                      mid_type='num', mid_value=0.5, mid_color='FFFFFFFF',
                      end_type='num', end_value=1, end_color='FF63BE7B')

def translate(deckname):
    if deckname in deckname_dict.keys():
        return deckname_dict[deckname]
    else:
        return deckname

def deck_merge(deckname):
    if deckname in deckmerge_dict.keys():
        return deckmerge_dict[deckname]
    else:
        return deckname

if __name__ == '__main__':
    # 0 for unlimited, 1 for rotation
    is_rot = 0
    rotation = ['无限', '指定']
    raw_filename = '1月第三周{}'.format(rotation[is_rot])
    processed_filename = raw_filename + '处理版.xlsx'
    raw_filename = raw_filename + '.xlsx'
    report = openpyxl.load_workbook(raw_filename)
    processed = openpyxl.Workbook()
    rp = report['{worksheet}']
    sheet1 = processed.active
    sheet1.title = '中文版'
    sheet2 = processed.create_sheet('英文版')
    sheet3 = processed.create_sheet('胜率方阵')
    decklist = {}
    
    for i in range(1, rp.max_column+1):
        if rp.cell(1, i).value == None:
            continue
        decks = rp.cell(1, i).value.split('.')
        if len(decks)>1 and deck_merge(decks[1]) not in decklist.keys():
            decklist[deck_merge(decks[1])] = len(decklist)
    
    print(decklist)
    # total, wins, 1st total, 1st wins in each matchup
    data = np.zeros((len(decklist), len(decklist), 4))
    
    for (deck, num) in decklist.items():
        sheet3.cell(num+2, 1).value = deck
        sheet3.cell(1, num+2).value = deck
    
    for i in range(1, rp.max_column+1):
        if rp.cell(1, i).value == None:
            continue
        decks = rp.cell(1, i).value.split('.')
        if len(decks) == 2:
            base_row = 7 * decklist[deck_merge(decks[1])]
            sheet1.cell(base_row+1, 1).value = deck_merge(decks[1])
            sheet1.cell(base_row+2, 2).value = '对局职业'
            sheet1.cell(base_row+3, 1).value = '总对局数'
            sheet1.cell(base_row+4, 1).value = '胜场'
            sheet1.cell(base_row+5, 1).value = '先手胜率'
            sheet1.cell(base_row+6, 1).value = '后手胜率'
            sheet1.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
            sheet1.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
            
            sheet2.cell(base_row+1, 1).value = translate(decks[1])
            sheet2.cell(base_row+2, 2).value = 'Opponent Deck'
            sheet2.cell(base_row+3, 1).value = 'Total Matches'
            sheet2.cell(base_row+4, 1).value = 'Wins'
            sheet2.cell(base_row+5, 1).value = 'Going 1st Winrate'
            sheet2.cell(base_row+6, 1).value = 'Going 2nd Winrate'
            sheet2.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
            sheet2.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
        elif len(decks) == 3:
            matches = rp.cell(2, i).value
            wins = rp.cell(3, i).value
            first = rp.cell(4, i).value
            first_wins = rp.cell(5, i).value
            if matches == None:
                matches = '0'
            if wins == None:
                wins = '0'
            if first == None:
                first = '0'
            if first_wins == None:
                first_wins = '0'
            matches = int(matches)
            wins = int(wins)
            first = int(first)
            first_wins = int(first_wins)
            
            data[decklist[deck_merge(decks[1])]][decklist[deck_merge(decks[2])]][0] += matches
            data[decklist[deck_merge(decks[1])]][decklist[deck_merge(decks[2])]][1] += wins
            data[decklist[deck_merge(decks[1])]][decklist[deck_merge(decks[2])]][2] += first
            data[decklist[deck_merge(decks[1])]][decklist[deck_merge(decks[2])]][3] += first_wins
    
    col = 0
    for (deck1, i) in decklist.items():
        col = 3
        base_row = 7 * i
        first = 0
        first_wins = 0
        for (deck2, j) in decklist.items():
            if data[i][j][0] > 0:
                first += data[i][j][2]
                first_wins += data[i][j][3]
                sheet1.cell(base_row+2, col).value = deck2
                sheet1.cell(base_row+3, col).value = data[i][j][0]
                sheet1.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet1.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet1.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet1.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet1.cell(base_row+6, col).number_format = '0.00%'
                
                sheet2.cell(base_row+2, col).value = translate(deck2)
                sheet2.cell(base_row+3, col).value = data[i][j][0]
                sheet2.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet2.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet2.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet2.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet2.cell(base_row+6, col).number_format = '0.00%'
                
                sheet3.cell(i+2, j+2).value = '={}/{}'.format(data[i][j][1], data[i][j][0])
                sheet3.cell(i+2, j+2).number_format = '0.00%'
                col += 1
        sheet1.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet1.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        sheet2.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet2.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        
        sheet1.cell(base_row+5, 2).number_format = '0.00%'
        sheet1.cell(base_row+6, 2).number_format = '0.00%'
        sheet2.cell(base_row+5, 2).number_format = '0.00%'
        sheet2.cell(base_row+6, 2).number_format = '0.00%'
    
    sheet3.conditional_formatting.add('A1:{}{}'.format(get_column_letter(1+len(decklist)), 1+len(decklist)), rule)
    processed.save(processed_filename)