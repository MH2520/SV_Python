import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.cell import get_column_letter
import numpy as np
from deckname_dict import deckname_dict
from deckmerge_dict import rot_decktrans, ult_decktrans, rot_deckmerge, ult_deckmerge
rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFF8696B',
                      mid_type='num', mid_value=0.5, mid_color='FFFFFFFF',
                      end_type='num', end_value=1, end_color='FF63BE7B')

web_rot = ['ult', 'rot']
app_rot = ['无限', '指定']

def translate(deck):
    if deck in deckname_dict.keys():
        return deckname_dict[deck]
    else:
        return deck

def merge(deck, is_rot):
    if is_rot == 1:
        if deck in rot_deckmerge.keys():
            return rot_deckmerge[deck]
        else:
            return deck
    else:
        if deck in ult_deckmerge.keys():
            return ult_deckmerge[deck]
        else:
            return deck

# mat_thres: minimum number of matches for a deck to be included in winrate matrix
# is_rot: 0 for unlimited, 1 for rotation
def main(webname, appname, deckname='{}_deck', mat_thres=5, is_rot=0):
    deck_filename = deckname.format(web_rot[is_rot])
    web_filename = webname.format(web_rot[is_rot])
    app_filename = appname.format(app_rot[is_rot])
    decksheet = openpyxl.load_workbook(deck_filename+'.xlsx')[deck_filename]
    web_data = openpyxl.load_workbook(web_filename+'.xlsx')['{}_match'.format(web_rot[is_rot])]
    app_data = openpyxl.load_workbook(app_filename+'.xlsx')['{worksheet}']
    
    processed_filename = app_filename + '综合处理版.xlsx'
    processed = openpyxl.Workbook()
    sheet1 = processed.active
    sheet1.title = '中文版'
    sheet2 = processed.create_sheet('英文版')
    sheet3 = processed.create_sheet('胜率方阵')
    decklist = {}
    deckid = {}
    
    for i in range(2, decksheet.max_row+1):
        decklist[decksheet.cell(i, 3).value] = i-2
        deckid[int(decksheet.cell(i, 1).value)] = decksheet.cell(i, 3).value
        
    print(decklist)
    # print(deckid)
    
    # total, wins, 1st total, 1st wins in each matchup
    data = np.zeros((len(decklist), len(decklist), 4))
    
    # reading web data
    for i in range(2, web_data.max_row+1):
        if web_data.cell(i, 1).value not in deckid.keys():
            continue
        deck1 = decklist[merge(deckid[web_data.cell(i, 1).value], is_rot)]
        deck2 = decklist[merge(deckid[web_data.cell(i, 2).value], is_rot)]
        first = web_data.cell(i, 3).value
        win = web_data.cell(i, 4).value
        data[deck1][deck2][0] += 1
        data[deck2][deck1][0] += 1
        if first == 't':
            data[deck1][deck2][2] += 1
            if win == 't':
                data[deck1][deck2][3] += 1
        elif first == 'f':
            data[deck2][deck1][2] += 1
            if win == 'f':
                data[deck2][deck1][3] += 1
        if win == 't':
            data[deck1][deck2][1] += 1
        elif win == 'f':
            data[deck2][deck1][1] += 1
    
    # reading app data
    for i in range(1, app_data.max_column+1):
        if app_data.cell(1, i).value == None:
            continue
        decks = app_data.cell(1, i).value.split('.')
        if len(decks) == 3:
            deck1 = decks[1]
            deck2 = decks[2]
            if is_rot == 1:
                if deck1 not in decklist.keys():
                    deck1 = rot_decktrans[deck1]
                if deck2 not in decklist.keys():
                    deck2 = rot_decktrans[deck2]
            else:
                if deck1 not in decklist.keys():
                    deck1 = ult_decktrans[deck1]
                if deck2 not in decklist.keys():
                    deck2 = ult_decktrans[deck2]
            deck1 = decklist[merge(deck1, is_rot)]
            deck2 = decklist[merge(deck2, is_rot)]
            matches = app_data.cell(2, i).value
            wins = app_data.cell(3, i).value
            first = app_data.cell(4, i).value
            first_wins = app_data.cell(5, i).value
            if matches == None:
                matches = '0'
            if wins == None:
                wins = '0'
            if first == None:
                first = '0'
            if first_wins == None:
                first_wins = '0'
            matches = int(matches)
            wins = int(wins)
            first = int(first)
            first_wins = int(first_wins)
            
            # verifying data
            if matches < wins or matches < first or first < first_wins or wins < first_wins:
                print('App data error at column {}'.format(get_column_letter(i)))
            
            data[deck1][deck2][0] += matches
            data[deck1][deck2][1] += wins
            data[deck1][deck2][2] += first
            data[deck1][deck2][3] += first_wins
    
    # verifying data
    if (data[:,:,0] != data[:,:,0].transpose()).any()\
    or (data[:,:,0] != data[:,:,1] + data[:,:,1].transpose()).any()\
    or (data[:,:,0] != data[:,:,2] + data[:,:,2].transpose()).any()\
    or (data[:,:,0] < data[:,:,1]).any()\
    or (data[:,:,0] < data[:,:,2]).any()\
    or (data[:,:,1] < data[:,:,3]).any()\
    or (data[:,:,2] < data[:,:,3]).any():
        print('Data error')
    
    # sorting decklist by number of matches, from high to low
    ord_decklist = list(decklist.items())
    ord_decklist.sort(key=(lambda x: np.sum(data, 1)[x[1]][0]), reverse=True)
    
    # writing labels for winrate matrix
    mat_row = 2
    mat_deckindex = {}
    for (deck, i) in ord_decklist:
        if np.sum(data, 1)[i][0] >= mat_thres:
            sheet3.cell(mat_row, 1).value = deck
            sheet3.cell(1, mat_row).value = deck
            mat_deckindex[i] = mat_row
            mat_row += 1
    
    # writing data
    col = 0
    base_row = -7
    for (deck1, i) in ord_decklist:
        if np.sum(data, 1)[i][0] == 0:
            continue
        col = 3
        base_row += 7
        
        first = 0
        first_wins = 0
        
        sheet1.cell(base_row+1, 1).value = deck1
        sheet1.cell(base_row+2, 2).value = '对局职业'
        sheet1.cell(base_row+3, 1).value = '总对局数'
        sheet1.cell(base_row+4, 1).value = '胜场'
        sheet1.cell(base_row+5, 1).value = '先手胜率'
        sheet1.cell(base_row+6, 1).value = '后手胜率'
        sheet1.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
        sheet1.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
        
        sheet2.cell(base_row+1, 1).value = translate(deck1)
        sheet2.cell(base_row+2, 2).value = 'Opponent Deck'
        sheet2.cell(base_row+3, 1).value = 'Total Matches'
        sheet2.cell(base_row+4, 1).value = 'Wins'
        sheet2.cell(base_row+5, 1).value = 'Going 1st Winrate'
        sheet2.cell(base_row+6, 1).value = 'Going 2nd Winrate'
        sheet2.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
        sheet2.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
        
        for (deck2, j) in ord_decklist:
            if data[i][j][0] > 0:
                first += data[i][j][2]
                first_wins += data[i][j][3]
                sheet1.cell(base_row+2, col).value = deck2
                sheet1.cell(base_row+3, col).value = data[i][j][0]
                sheet1.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet1.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet1.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet1.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet1.cell(base_row+6, col).number_format = '0.00%'
                
                sheet2.cell(base_row+2, col).value = translate(deck2)
                sheet2.cell(base_row+3, col).value = data[i][j][0]
                sheet2.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet2.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet2.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet2.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet2.cell(base_row+6, col).number_format = '0.00%'
                
                if np.sum(data, 1)[i][0] >= mat_thres and np.sum(data, 1)[j][0] >= mat_thres:
                    sheet3.cell(mat_deckindex[i], mat_deckindex[j]).value = '={}/{}'.format(data[i][j][1], data[i][j][0])
                    sheet3.cell(mat_deckindex[i], mat_deckindex[j]).number_format = '0.00%'
                col += 1
        
        sheet1.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet1.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        sheet2.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet2.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        
        sheet1.cell(base_row+5, 2).number_format = '0.00%'
        sheet1.cell(base_row+6, 2).number_format = '0.00%'
        sheet2.cell(base_row+5, 2).number_format = '0.00%'
        sheet2.cell(base_row+6, 2).number_format = '0.00%'
    
    sheet3.conditional_formatting.add('A1:{}{}'.format(get_column_letter(1+len(decklist)), 1+len(decklist)), rule)
    processed.save(processed_filename)
    
if __name__ == '__main__':
    main(webname='1-18~1-26{}', appname='1月第四周{}', is_rot=0)