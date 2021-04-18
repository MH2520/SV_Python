import util
import openpyxl
from deck_classify_dict import rot_deck, unl_deck
from altart_dict import alt2std

classes = ['精灵', '皇室', '法师', '龙族', '死灵', '吸血鬼', '主教', '复仇者']

# Counts the number of each card and checks legality of the deck
# Standardizes all alt-art cards in the deck
# Input: hash of a deck, which can be obtained by util.get_hash
# Output: class number and a dictionary mapping card id to the number of copies in deck
def count_cards(hs):
    bd = util.hash_breakdown(hs)
    if len(bd) != 42:
        return (0, 'Illegal Deck')
    cl = bd[1]
    cnt = {}
    for i in range(2, len(bd)):
        if bd[i]//100000%10 != cl and bd[i]//100000%10 != 0:
            return (0, 'Illegal Deck')
        cid = bd[i]
        if bd[i] in alt2std:
            cid = alt2std[bd[i]]
        if cid not in cnt:
            cnt[cid] = 1
        else:
            cnt[cid] += 1
    return (cl, cnt)

def classify(mode, cl, cnt):
    if mode == 0: # rotation
        deckcodes = rot_deck
    else: # unlimited
        deckcodes = unl_deck
    
    for (deck, crit) in deckcodes.items():
        is_deck = True
        if cl != crit[0]:
            continue
        for i in range((len(crit)-1)//3):
            cid = crit[3*i+1]
            op = crit[3*i+2]
            num = crit[3*i+3]
            if op == 0: # >=
                if cid not in cnt or cnt[cid] < num:
                    is_deck = False
                    break
            elif op == 1: # <=
                if cid in cnt and cnt[cid] > num:
                    is_deck = False
                    break
        if is_deck:
            return deck
    return '其它' + classes[cl-1]

if __name__ == '__main__':
    sheet = openpyxl.load_workbook('decklist.xlsx')['Sheet1']
    for i in range(8):
        output = '{}: '.format(sheet.cell(i+1, 1).value)
        for j in range(3):
            (cl, cnt) = count_cards(util.get_hash(sheet.cell(i+1, j+2).value))
            output += classify(0, cl, cnt)
            if j != 3:
                output += ', '
        print(output)