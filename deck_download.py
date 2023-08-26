import openpyxl
import deck_image
import time
import argparse

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Download deck images of decklists in decklist.xlsx')
    parser.add_argument('player', nargs='?', default=8, help='Number of players')
    parser.add_argument('deck', nargs='?', default=3, help='Number of decks per player')
    parser.add_argument('month', nargs='?', default=time.localtime().tm_mon, help='Month of competition')
    parser.add_argument('sheet', nargs='?', default='decklist.xlsx', help='Decklist spreadsheet file')
    args = parser.parse_args()
    player = args.player
    deck = args.deck
    month = args.month
    sheet = openpyxl.load_workbook(args.sheet)['Sheet1']
    names = []
    nums = []
    links = []
    for i in range(player):
        for j in range(1, deck+1):
            names.append('{}_'.format(month)+str(player-i)+'_'+str(sheet.cell(1+i, 1).value))
            nums.append(4-j)
            links.append(sheet.cell(1+i, 3+j).value)
    deck_image.batch_download_deck(names, nums, links)