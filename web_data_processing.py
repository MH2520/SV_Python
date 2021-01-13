import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.cell import get_column_letter
import xlrd
import numpy as np
from deckname_dict import deckname_dict
rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFF8696B',
                      mid_type='num', mid_value=0.5, mid_color='FFFFFFFF',
                      end_type='num', end_value=1, end_color='FF63BE7B')
def translate(deckname):
    if deckname in deckname_dict.keys():
        return deckname_dict[deckname]
    else:
        return deckname

if __name__ == '__main__':
    # minimum number of matches for a deck to be included in winrate matrix
    mat_thres = 5
    
    # 0 for unlimited, 1 for rotation
    is_rot = 1
    rotation = ['ult', 'rot']
    decksheet_filename = '{}_deck'.format(rotation[is_rot])
    decksheet = openpyxl.load_workbook(decksheet_filename+'.xlsx')[decksheet_filename]
    # decksheet = xlrd.open_workbook(decksheet_filename+'.xlsx').sheet_by_index(0)
    datasheet_filename = '{}_match'.format(rotation[is_rot])
    datasheet = openpyxl.load_workbook(datasheet_filename+'.xlsx')[datasheet_filename]
    # datasheet = xlrd.open_workbook(datasheet_filename+'.xlsx').sheet_by_index(0)
    
    processed_filename = datasheet_filename + '处理版.xlsx'
    processed = openpyxl.Workbook()
    sheet1 = processed.active
    sheet1.title = '中文版'
    sheet2 = processed.create_sheet('英文版')
    sheet3 = processed.create_sheet('胜率方阵')
    decklist = {}
    deckid = {}
    
    # for i in range(1, decksheet.nrows):
        # decklist[decksheet.cell_value(i, 2)] = i-1
        # deckid[int(decksheet.cell_value(i, 0))] = decksheet.cell_value(i, 2)
    
    for i in range(2, decksheet.max_row+1):
        decklist[decksheet.cell(i, 3).value] = i-2
        deckid[int(decksheet.cell(i, 1).value)] = decksheet.cell(i, 3).value
        
    print(decklist)
    print(deckid)
    
    # total, wins, 1st total, 1st wins in each matchup
    data = np.zeros((len(decklist), len(decklist), 4))
    
    # writing data into data matrix
    # for i in range(1, datasheet.nrows):
    #     deck1 = decklist[deckid[datasheet.cell_value(i, 0)]]
    #     deck2 = decklist[deckid[datasheet.cell_value(i, 1)]]
    #     first = datasheet.cell_value(i, 2)
    #     win = datasheet.cell_value(i, 3)
    #     data[deck1][deck2][0] += 1
    #     data[deck2][deck1][0] += 1
    #     if first == 't':
    #         data[deck1][deck2][2] += 1
    #         if win == 't':
    #             data[deck1][deck2][3] += 1
    #     elif first == 'f':
    #         data[deck2][deck1][2] += 1
    #         if win == 'f':
    #             data[deck2][deck1][3] += 1
    #     if win == 't':
    #         data[deck1][deck2][1] += 1
    #     elif win == 'f':
    #         data[deck2][deck1][1] += 1
    for i in range(2, datasheet.max_row+1):
        if datasheet.cell(i, 1).value not in deckid.keys():
            continue
        deck1 = decklist[deckid[datasheet.cell(i, 1).value]]
        deck2 = decklist[deckid[datasheet.cell(i, 2).value]]
        first = datasheet.cell(i, 3).value
        win = datasheet.cell(i, 4).value
        data[deck1][deck2][0] += 1
        data[deck2][deck1][0] += 1
        if first == 't':
            data[deck1][deck2][2] += 1
            if win == 't':
                data[deck1][deck2][3] += 1
        elif first == 'f':
            data[deck2][deck1][2] += 1
            if win == 'f':
                data[deck2][deck1][3] += 1
        if win == 't':
            data[deck1][deck2][1] += 1
        elif win == 'f':
            data[deck2][deck1][1] += 1
    
    mat_row = 2
    mat_deckindex = {}
    for (deck, i) in decklist.items():
        if np.sum(data, 1)[i][0] >= mat_thres:
            sheet3.cell(mat_row, 1).value = deck
            sheet3.cell(1, mat_row).value = deck
            mat_deckindex[i] = mat_row
            mat_row += 1
    
    col = 0
    base_row = -7
    for (deck1, i) in decklist.items():
        if np.sum(data, 1)[i][0] == 0:
            continue
        col = 3
        base_row += 7
        
        first = 0
        first_wins = 0
        
        sheet1.cell(base_row+1, 1).value = deck1
        sheet1.cell(base_row+2, 2).value = '对局职业'
        sheet1.cell(base_row+3, 1).value = '总对局数'
        sheet1.cell(base_row+4, 1).value = '胜场'
        sheet1.cell(base_row+5, 1).value = '先手胜率'
        sheet1.cell(base_row+6, 1).value = '后手胜率'
        sheet1.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
        sheet1.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
        
        sheet2.cell(base_row+1, 1).value = translate(deck1)
        sheet2.cell(base_row+2, 2).value = 'Opponent Deck'
        sheet2.cell(base_row+3, 1).value = 'Total Matches'
        sheet2.cell(base_row+4, 1).value = 'Wins'
        sheet2.cell(base_row+5, 1).value = 'Going 1st Winrate'
        sheet2.cell(base_row+6, 1).value = 'Going 2nd Winrate'
        sheet2.cell(base_row+3, 2).value = '=SUM(C{}:{}{})'.format(base_row+3, get_column_letter(3+len(decklist)), base_row+3)
        sheet2.cell(base_row+4, 2).value = '=SUM(C{}:{}{})'.format(base_row+4, get_column_letter(3+len(decklist)), base_row+4)
        
        for (deck2, j) in decklist.items():
            if data[i][j][0] > 0:
                first += data[i][j][2]
                first_wins += data[i][j][3]
                sheet1.cell(base_row+2, col).value = deck2
                sheet1.cell(base_row+3, col).value = data[i][j][0]
                sheet1.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet1.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet1.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet1.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet1.cell(base_row+6, col).number_format = '0.00%'
                
                sheet2.cell(base_row+2, col).value = translate(deck2)
                sheet2.cell(base_row+3, col).value = data[i][j][0]
                sheet2.cell(base_row+4, col).value = data[i][j][1]
                if data[i][j][2] != 0:
                    sheet2.cell(base_row+5, col).value = '={}/{}'.format(data[i][j][3], data[i][j][2])
                    sheet2.cell(base_row+5, col).number_format = '0.00%'
                if data[i][j][0] - data[i][j][2] != 0: 
                    sheet2.cell(base_row+6, col).value = '={}/{}'.format(data[i][j][1] - data[i][j][3], data[i][j][0] - data[i][j][2])
                    sheet2.cell(base_row+6, col).number_format = '0.00%'
                
                if np.sum(data, 1)[i][0] >= mat_thres and np.sum(data, 1)[j][0] >= mat_thres:
                    sheet3.cell(mat_deckindex[i], mat_deckindex[j]).value = '={}/{}'.format(data[i][j][1], data[i][j][0])
                    sheet3.cell(mat_deckindex[i], mat_deckindex[j]).number_format = '0.00%'
                col += 1
        
        sheet1.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet1.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        sheet2.cell(base_row+5, 2).value = '={}/{}'.format(first_wins, first)
        sheet2.cell(base_row+6, 2).value = '=(B{}-{})/(B{}-{})'.format(base_row+4, first_wins, base_row+3, first)
        
        sheet1.cell(base_row+5, 2).number_format = '0.00%'
        sheet1.cell(base_row+6, 2).number_format = '0.00%'
        sheet2.cell(base_row+5, 2).number_format = '0.00%'
        sheet2.cell(base_row+6, 2).number_format = '0.00%'
    
    sheet3.conditional_formatting.add('A1:{}{}'.format(get_column_letter(1+len(decklist)), 1+len(decklist)), rule)
    processed.save(processed_filename)